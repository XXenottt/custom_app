import re
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

WORKBOOK_PATH = Path('/home/zxh/custom_app/src/EU LTL DELIVERY RATE_H1_2026.xlsx')
OUT_XLSX = Path('/home/zxh/custom_app/src/city_value_mapping.xlsx')
OUT_PDF = Path('/home/zxh/custom_app/src/city_value_mapping.pdf')
OVERRIDES_PATH = Path('/home/zxh/custom_app/src/manual_overrides.json')


@dataclass
class CityRule:
    country: str
    city: str
    postal_area: str


CITY_RULES: List[CityRule] = [
    CityRule('AT', 'Vienna', '1xxx'), CityRule('AT', 'Graz', '8xxx'), CityRule('AT', 'Salzburg', '5xxx'), CityRule('AT', 'Innsbruck', '6xxx'), CityRule('AT', 'Linz', '4xxx'),
    CityRule('BA', 'Sarajevo', 'Sarajevo'), CityRule('BA', 'Banja Luka', 'Banje Luka'), CityRule('BA', 'Mostar', 'Mostar'),
    CityRule('BE', 'Brussels', '1xxx'), CityRule('BE', 'Antwerp', '2xxx'), CityRule('BE', 'Ghent', '9xxx'), CityRule('BE', 'Liège', '4xxx'), CityRule('BE', 'Bruges', '8xxx'),
    CityRule('BG', 'Sofia', '1xxx'), CityRule('BG', 'Plovdiv', '4xxx'), CityRule('BG', 'Varna', '9xxx'), CityRule('BG', 'Burgas', '8xxx'),
    CityRule('CH', 'Zurich', '8xxx'), CityRule('CH', 'Geneva', '1xxx'), CityRule('CH', 'Basel', '4xxx'), CityRule('CH', 'Bern', '3xxx'), CityRule('CH', 'Lausanne', '1xxx'),
    CityRule('CZ', 'Prague', '1xx'), CityRule('CZ', 'Brno', '6xx'), CityRule('CZ', 'Ostrava', '7xx'), CityRule('CZ', 'Pilsen', '3xx'),
    CityRule('DE', 'Berlin', '1xxxx'), CityRule('DE', 'Hamburg', '2xxxx'), CityRule('DE', 'Munich', '8xxxx'), CityRule('DE', 'Cologne', '5xxxx'), CityRule('DE', 'Frankfurt', '6xxxx'), CityRule('DE', 'Stuttgart', '7xxxx'), CityRule('DE', 'Leipzig', '0xxxx'), CityRule('DE', 'Düsseldorf', '4xxxx'),
    CityRule('DK', 'Copenhagen', '1xxx-2xxx'), CityRule('DK', 'Aarhus', '8xxx'), CityRule('DK', 'Odense', '5xxx'), CityRule('DK', 'Aalborg', '9xxx'),
    CityRule('EE', 'Tallinn', '1xxxx'), CityRule('EE', 'Tartu', '5xxxx'), CityRule('EE', 'Narva', '2xxxx'),
    CityRule('ES', 'Madrid', '28xxx'), CityRule('ES', 'Barcelona', '08xxx'), CityRule('ES', 'Valencia', '46xxx'), CityRule('ES', 'Seville', '41xxx'), CityRule('ES', 'Bilbao', '48xxx'), CityRule('ES', 'Zaragoza', '50xxx'),
    CityRule('FI', 'Helsinki', '00xxx'), CityRule('FI', 'Espoo', '02xxx'), CityRule('FI', 'Tampere', '33xxx'), CityRule('FI', 'Turku', '20xxx'),
    CityRule('FR', 'Paris', '75xxx'), CityRule('FR', 'Marseille', '130xx'), CityRule('FR', 'Lyon', '690xx'), CityRule('FR', 'Toulouse', '310xx'), CityRule('FR', 'Bordeaux', '330xx'), CityRule('FR', 'Lille', '590xx'), CityRule('FR', 'Nantes', '440xx'),
    CityRule('HR', 'Zagreb', '1xxxx'), CityRule('HR', 'Split', '2xxxx'), CityRule('HR', 'Rijeka', '5xxxx'),
    CityRule('HU', 'Budapest', '1xxx'), CityRule('HU', 'Debrecen', '4xxx'), CityRule('HU', 'Pécs', '7xxx'), CityRule('HU', 'Miskolc', '3xxx'),
    CityRule('IE', 'Dublin', 'D01-D24'), CityRule('IE', 'Cork', 'T12'), CityRule('IE', 'Galway', 'H91'), CityRule('IE', 'Limerick', 'V94'),
    CityRule('IT', 'Rome', '001xx'), CityRule('IT', 'Milan', '201xx'), CityRule('IT', 'Naples', '801xx'), CityRule('IT', 'Turin', '101xx'), CityRule('IT', 'Bologna', '401xx'), CityRule('IT', 'Florence', '501xx'), CityRule('IT', 'Venice', '301xx'),
    CityRule('LT', 'Vilnius', '0xxxx'), CityRule('LT', 'Kaunas', '4xxxx'), CityRule('LT', 'Klaipėda', '9xxxx'),
    CityRule('LU', 'Luxembourg', 'Lxxx'), CityRule('LU', 'Esch-sur-Alzette', 'Lxxx'),
    CityRule('LV', 'Riga', 'LV-1xxx'), CityRule('LV', 'Daugavpils', 'LV-5xxx'),
    CityRule('NO', 'Oslo', '0xxx'), CityRule('NO', 'Bergen', '5xxx'), CityRule('NO', 'Trondheim', '7xxx'), CityRule('NO', 'Stavanger', '4xxx'),
    CityRule('PL', 'Warsaw', '00-xx-03-xx'), CityRule('PL', 'Kraków', '30-xx-31-xx'), CityRule('PL', 'Łódź', '90-xx-94-xx'), CityRule('PL', 'Wrocław', '50-xx-54-xx'), CityRule('PL', 'Poznań', '60-xx-61-xx'),
    CityRule('PT', 'Lisbon', '1xxx-xxx'), CityRule('PT', 'Porto', '4xxx-xxx'), CityRule('PT', 'Braga', '47xx'),
    CityRule('RO', 'Bucharest', '0xxxx'), CityRule('RO', 'Cluj-Napoca', '4xxxx'), CityRule('RO', 'Timișoara', '3xxxx'), CityRule('RO', 'Iași', '7xxxx'),
    CityRule('RS', 'Belgrade', 'Belgardo'), CityRule('RS', 'Novi Sad', '21000'), CityRule('RS', 'Niš', '18000'),
    CityRule('SE', 'Stockholm', '1xxxx'), CityRule('SE', 'Gothenburg', '4xxxx'), CityRule('SE', 'Malmö', '2xxxx'), CityRule('SE', 'Uppsala', '7xxxx'),
    CityRule('SI', 'Ljubljana', '1xxx'), CityRule('SI', 'Maribor', '2xxx'),
    CityRule('SK', 'Bratislava', '8xxxx'), CityRule('SK', 'Košice', '0xxxx'),
    CityRule('UK', 'London', 'E/EC/N/NW/SE/SW/W/WC'), CityRule('UK', 'Manchester', 'M'), CityRule('UK', 'Birmingham', 'B'), CityRule('UK', 'Leeds', 'LS'), CityRule('UK', 'Glasgow', 'G'), CityRule('UK', 'Edinburgh', 'EH'), CityRule('UK', 'Bristol', 'BS'), CityRule('UK', 'Liverpool', 'L'),
]


def norm(s: str) -> str:
    return re.sub(r'[^A-Z0-9]', '', str(s).upper())


def parse_float(v) -> Optional[float]:
    if v is None:
        return None
    txt = str(v).strip().replace(',', '.')
    try:
        return float(txt)
    except ValueError:
        return None


def find_economy_and_tier(ws):
    eco_row = None
    tier_row = None
    header_labels = {'economy', 'distribution'}
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or '').strip().lower() in header_labels:
            eco_row = r
            break
    if eco_row is None:
        return None, None

    for r in range(eco_row + 1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or '').strip().lower()
        if a in header_labels:
            break
        if a.startswith('751-1000'):
            tier_row = r
            break
    return eco_row, tier_row


def economy_map(ws, eco_row: int, tier_row: int) -> Dict[str, float]:
    m: Dict[str, float] = {}
    for c in range(1, ws.max_column + 1):
        code = ws.cell(eco_row, c).value
        if code is None:
            continue
        code = str(code).strip()
        if not code or code.lower() in {'economy', 'distribution'}:
            continue
        val = parse_float(ws.cell(tier_row, c).value)
        if val is None:
            continue
        m[code] = val
    return m


DEFAULT_MANUAL_OVERRIDES: Dict[Tuple[str, str], List[str]] = {
    ('DE', 'Leipzig'): ['4'],
    ('RO', 'Bucharest'): ['1', '2', '3', '4', '5', '6'],
    ('SK', 'Košice'): ['4'],
    ('LU', 'Luxembourg'): ['10'],
    ('LU', 'Esch-sur-Alzette'): ['10'],
}


def load_manual_overrides(path: Path) -> Dict[Tuple[str, str], List[str]]:
    if not path.exists():
        return dict(DEFAULT_MANUAL_OVERRIDES)

    try:
        raw = json.loads(path.read_text(encoding='utf-8'))
    except (json.JSONDecodeError, OSError):
        return dict(DEFAULT_MANUAL_OVERRIDES)

    out: Dict[Tuple[str, str], List[str]] = {}
    for key, value in raw.items():
        if '|' not in key or not isinstance(value, list):
            continue
        country, city = key.split('|', 1)
        country = country.strip()
        city = city.strip()
        codes = [str(v).strip() for v in value if str(v).strip()]
        if country and city and codes:
            out[(country, city)] = codes

    if not out:
        return dict(DEFAULT_MANUAL_OVERRIDES)
    return out


def match_codes(pattern: str, econ_codes: List[str]) -> Set[str]:
    out: Set[str] = set()
    p = pattern.strip().upper()

    # UK style multiple prefixes
    if '/' in p and not any(ch.isdigit() for ch in p):
        allowed = {norm(x) for x in p.split('/') if x.strip()}
        for code in econ_codes:
            if norm(code.replace('*', '')) in allowed:
                out.add(code)
        return out

    # textual direct match (BA/RS style)
    if any(ch.isalpha() for ch in p) and 'X' not in p and '-' not in p and '/' not in p:
        pn = norm(p)
        for code in econ_codes:
            cn = norm(code.replace('*', ''))
            if pn and (pn in cn or cn in pn):
                out.add(code)
        if out:
            return out

    # IE range D01-D24
    m_ie = re.fullmatch(r'([A-Z])(\d{2})-([A-Z])(\d{2})', p)
    if m_ie and m_ie.group(1) == m_ie.group(3):
        letter = m_ie.group(1)
        lo = int(m_ie.group(2))
        hi = int(m_ie.group(4))
        for code in econ_codes:
            cn = norm(code.replace('*', ''))
            m = re.fullmatch(r'([A-Z])(\d{2,3})', cn)
            if not m or m.group(1) != letter:
                continue
            n = int(m.group(2)[:2])
            if lo <= n <= hi:
                out.add(code)
        return out

    # generic numeric wildcard/range
    parts = re.findall(r'[A-Z]*\d+[A-Z]*', p)
    if '-' in p and len(parts) >= 2:
        lo = re.sub(r'\D', '', parts[0])
        hi = re.sub(r'\D', '', parts[1])
        if lo and hi:
            try:
                lo_i = int(lo)
                hi_i = int(hi)
                width = max(len(lo), len(hi))
                seeds = [str(i).zfill(width) for i in range(lo_i, hi_i + 1)]
            except ValueError:
                seeds = [lo, hi]
        else:
            seeds = []
    else:
        seeds = []
        m = re.match(r'([A-Z-]*)(\d+)', p)
        if m:
            seeds.append(m.group(2))

    # pure alpha prefix (UK single like M/LS)
    if not seeds and re.fullmatch(r'[A-Z]+', p):
        for code in econ_codes:
            if norm(code.replace('*', '')) == p:
                out.add(code)
        return out

    def code_ok(seed: str, econ: str) -> bool:
        ec_raw = norm(econ.replace('*', ''))
        if not ec_raw:
            return False

        # economy range e.g. 070-076*
        ec_range = re.match(r'^(\d+)-(\d+)$', norm(econ.replace('*', '').replace(' ', '').replace('--', '-')))
        if ec_range and seed.isdigit():
            width = len(ec_range.group(1))
            n = int(seed[:width].zfill(width))
            return int(ec_range.group(1)) <= n <= int(ec_range.group(2))

        seed2 = seed
        if seed2.startswith('0'):
            seed2_alt = seed2.lstrip('0') or '0'
        else:
            seed2_alt = seed2
        ec_alt = ec_raw.lstrip('0') or '0'
        return (
            seed2.startswith(ec_raw)
            or seed2_alt.startswith(ec_alt)
            or ec_raw.startswith(seed2)
            or ec_alt.startswith(seed2_alt)
        )

    for code in econ_codes:
        if any(code_ok(seed, code) for seed in seeds):
            out.add(code)

    return out


def build_pdf(rows: List[List[str]], out_pdf: Path) -> None:
    doc = SimpleDocTemplate(str(out_pdf), pagesize=landscape(A4))
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    doc.build([table])


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    manual_overrides = load_manual_overrides(OVERRIDES_PATH)

    result: List[List[str]] = []
    validations: List[List[str]] = []
    pdf_rows: List[List[str]] = [['Country', 'City', 'Matched abbreviations', 'Value(751-1000 max)']]

    for rule in CITY_RULES:
        if rule.country not in wb.sheetnames:
            result.append([rule.country, rule.city, rule.postal_area, '', '', 'sheet_not_found'])
            continue

        ws = wb[rule.country]
        eco_row, tier_row = find_economy_and_tier(ws)
        if eco_row is None or tier_row is None:
            # try manual override only if sheet exists but header not standard
            if (rule.country, rule.city) not in manual_overrides:
                result.append([rule.country, rule.city, rule.postal_area, '', '', 'economy_or_tier_missing'])
                continue

        emap = economy_map(ws, eco_row, tier_row) if eco_row and tier_row else {}
        codes = list(emap.keys())

        matched = match_codes(rule.postal_area, codes)
        status = 'ok'

        if (not matched or not emap) and (rule.country, rule.city) in manual_overrides:
            overrides = manual_overrides[(rule.country, rule.city)]
            matched = {c for c in overrides if c in emap} if emap else set(overrides)
            status = 'ok_manual_override' if matched else 'manual_override_failed'

        vals = [emap[c] for c in matched if c in emap]
        maxv = max(vals) if vals else None

        matched_text = ', '.join(sorted(matched))
        val_text = '' if maxv is None else f'{maxv:.2f}'
        if maxv is None:
            if status.startswith('ok_manual'):
                status = 'manual_override_no_value'
            else:
                status = 'no_match'

        result.append([rule.country, rule.city, rule.postal_area, matched_text, val_text, status])
        pdf_rows.append([rule.country, rule.city, matched_text, val_text])

        # validation: values must come from current sheet + tier row
        if maxv is not None:
            source_values = [emap[c] for c in matched if c in emap]
            source_max = max(source_values) if source_values else None
            ok = source_max == maxv
            validations.append([
                rule.country,
                rule.city,
                matched_text,
                f'{source_max:.2f}' if source_max is not None else '',
                val_text,
                'PASS' if ok else 'FAIL',
            ])

    out_wb = openpyxl.Workbook()
    ws_out = out_wb.active
    ws_out.title = 'city_values'
    ws_out.append(['Country', 'City', 'PostalAreaRule', 'MatchedAbbreviations', 'Value_751_1000_Max', 'Status'])
    for row in result:
        ws_out.append(row)

    ws_val = out_wb.create_sheet('validation')
    ws_val.append(['Country', 'City', 'MatchedAbbreviations', 'SourceMaxFromSheet', 'OutputValue', 'Check'])
    for row in validations:
        ws_val.append(row)
    out_wb.save(OUT_XLSX)

    build_pdf(pdf_rows, OUT_PDF)

    print(f'Generated: {OUT_XLSX}')
    print(f'Generated: {OUT_PDF}')


if __name__ == '__main__':
    main()
