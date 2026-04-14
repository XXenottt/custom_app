import os
import pandas as pd
import openpyxl
import re
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def process_files():
    inv_data = []
    pl_data = []
    
    files = sorted([f for f in os.listdir('.') if os.path.isfile(f)])
    
    inv_files = [f for f in files if "INV" in f]
    for file in inv_files:
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            ijk9_value = ws['I9'].value
            
            row = 19
            is_first_row_of_file = True
            while ws.cell(row=row, column=1).value is not None:
                # 提取原B到K列的数据（即第2到11列）
                # 按照逻辑：B, C, D, E, F, G, H(空), I, J, K
                # 你的需求：在A后插入空列(即新B列)，并将原K列移动到原H列位置，且删除原K列
                
                row_values = []
                # 原B到G列 (列索引 2 到 7)
                for col in range(2, 8):
                    row_values.append(ws.cell(row=row, column=col).value)
                
                # 原H列位置放入原K列的值 (列索引 11)
                row_values.append(ws.cell(row=row, column=11).value)
                
                # 原I到J列 (列索引 9 到 10)
                for col in range(9, 11):
                    row_values.append(ws.cell(row=row, column=col).value)
                
                # 构造最终行：[A列值, 空白列(新B), 原B, 原C, 原D, 原E, 原F, 原G, 原K(现H), 原I, 原J]
                a_column_val = ijk9_value if is_first_row_of_file else ""
                final_row = [a_column_val, ""] + row_values
                
                inv_data.append(final_row)
                is_first_row_of_file = False
                row += 1
        except Exception:
            continue

    pl_files = [f for f in files if "packing list" in f]
    for file in pl_files:
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            found = False
            for row in ws.iter_rows():
                for cell in row:
                    val = str(cell.value) if cell.value else ""
                    if "Total：" in val and "net weight(KG)：" in val:
                        cases = re.search(r"Total：(\d+)\s+CASES", val)
                        net_weight = re.search(r"net weight\(KG\)：([\d.]+)", val)
                        if cases and net_weight:
                            pl_data.append([net_weight.group(1), cases.group(1)])
                            found = True
                            break
                if found: break
        except Exception:
            continue

    with pd.ExcelWriter('output.xlsx') as writer:
        if inv_data:
            df_inv = pd.DataFrame(inv_data)
            df_inv.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
        
        if pl_data:
            df_pl = pd.DataFrame(pl_data)
            df_pl.to_excel(writer, sheet_name='Sheet2', index=False, header=False)

if __name__ == "__main__":
    process_files()